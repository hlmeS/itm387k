#! /usr/bin/env python2

"""

Title: Scraping Yelp
Author: Holm Smidt
Version: 0.9
Date: 11-11-2017

Overview:
* Simple script to use BeautifulSoup4 package for web scraping.
* The goal is to retriever Yelp Reviews.
* Basic html understanding is recommended.

* Code adapted from: https://goo.gl/pCF51R
"""

import urllib2
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import sys  # import sys package, if not already imported

# some encoding matters
reload(sys)
sys.setdefaultencoding('utf-8')

def get_yelp_review(url, many):
    page = urllib2.urlopen(url)
    soup = BeautifulSoup(page, 'html.parser')

    if many:
        # identify 'p', 'lang', 'en' from the html source code
        scrapes = soup.findAll('p', attrs={'lang': 'en'})
        output = []
        for scrape in scrapes:
            output.append(scrape.text)
    else:
        scrape = soup.find('p', attrs={'lang': 'en'})
        output = scrape.text

    return output

def retrieve_and_write_to_excel(filename, urls):

    wb = Workbook()
    dest_filename = 'yelp_data.xlsx'

    for i, url in enumerate(urls):

        # retrieve reviews, set many to True
        reviews = get_yelp_review(url, True)
        if i == 0:
            ws = wb.active
            ws.title = url.split("/biz/", 1)[1].split("?", 1)[0][:15]
        else :
            ws = wb.create_sheet(title= url.split("/biz/", 1)[1].split("?", 1)[0][:15] )

        ws['A1'] = "ID"
        ws['B1'] = "Review"
        for i, review in enumerate(reviews):
            _= ws.cell(column=1, row=i+2, value=i+1)
            _= ws.cell(column=2, row=i+2, value=unicode(str(review), 'utf-8'))

    wb.save(filename = dest_filename)

def run():

    coffeeshops = ['https://www.yelp.com/biz/island-vintage-coffee-honolulu-4?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/local-joe-honolulu?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/island-brew-coffeehouse-honolulu?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/kai-coffee-hawaii-honolulu?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/kona-coffee-purveyors-and-patisserie-honolulu-2?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/9bar-hnl-honolulu?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/the-curb-kaimuki-honolulu-2?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/morning-glass-coffee-cafe-honolulu-2?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/aloha-bakehouse-and-cafe-honolulu-2?osq=Coffee+%26+Tea',
                   'https://www.yelp.com/biz/brue-bar-honolulu-7?osq=Coffee+%26+Tea']

    retrieve_and_write_to_excel("yelp-coffeeshops.xlsx", coffeeshops)

run()
